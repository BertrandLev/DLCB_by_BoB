import pyqtgraph as pg
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from models.fit_results_model import ParameterTableModel
from utils.GPC_data import GPC
from utils import Flory_fit
from utils.Log_box import Log_box
from utils.Plot_box import Plot_box
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (QWidget, QGridLayout, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, 
                             QLineEdit, QSplitter, QTableView, QFileDialog, QGroupBox, QSpinBox,
                             QTextEdit)


class FloryFitTab(QWidget):
    
    def __init__(self) -> None:
        super().__init__()
        # Data
        self.data_GPC = GPC()
        self.param_model = ParameterTableModel()

        # Main Layout
        main_layout = QHBoxLayout(self)
        main_splitter = QSplitter()
        main_splitter.setOrientation(Qt.Orientation.Horizontal)
        main_layout.addWidget(main_splitter)

        # Left Part
        left_part = QWidget(main_splitter)
        left_layout = QVBoxLayout(left_part)
        # Add file widgets
        file_box = QGroupBox("File Selection")
        file_layout = QGridLayout(file_box)
        file_label = QLabel("File Name")
        self.file_entry = QLineEdit("")
        file_button = QPushButton("...")
        file_button.clicked.connect(self.openFileDialog)
        file_infoTitle = QLabel("Sample info:")
        self.file_info = QTextEdit()
        self.file_info.setFrameShape(QLabel.Shape.Box)
        self.file_info.setReadOnly(True)
        self.file_info.setFixedHeight(100)
        file_layout.addWidget(file_label,0,0)
        file_layout.addWidget(self.file_entry,1,0)
        file_layout.addWidget(file_button,1,1)
        file_layout.addWidget(file_infoTitle,2,0)
        file_layout.addWidget(self.file_info,3,0,1,2)
        # Add fitting option widgets
        fit_box = QGroupBox("Fit Options")
        fit_layout = QGridLayout(fit_box)
        fit_label = QLabel("Number of Flory")
        self.fit_entry = QSpinBox()
        self.fit_entry.setMinimumWidth(80)
        self.fit_entry.setValue(1)
        self.fit_entry.valueChanged.connect(self.floryNumberChange)
        fit_layout.addWidget(fit_label,0,0,)
        fit_layout.addWidget(self.fit_entry,0,1)
        fit_layout.setColumnStretch(0,1)
        # Add log widgets
        self.log = Log_box("Flory fit Log")
        # Left layout
        left_layout.addWidget(file_box,0)
        left_layout.addWidget(fit_box,1)
        left_layout.addWidget(self.log,2)
        left_layout.setStretch(1,0)
        left_layout.setStretch(2,1)

        # Right Part
        right_splitter = QSplitter(main_splitter)
        right_splitter.setOrientation(Qt.Orientation.Vertical)
        r_top_part = QWidget(right_splitter)
        r_top_layout = QVBoxLayout(r_top_part)
        self.plot_GPC = Plot_box(xlabel='Log M', ylabel='w')
        r_top_layout.addWidget(self.plot_GPC)
        r_bot_part = QWidget(right_splitter)
        r_bot_layout = QGridLayout(r_bot_part)
        self.equation_label = QTextEdit("Fitting equation : W(logM) = 2.3026 * M² * \u2211 [m<sub>i</sub> * \u03C4<sub>i</sub>² * exp(-\u03C4<sub>i</sub> * M)]")
        self.equation_label.setFixedHeight(40)
        self.equation_label.setReadOnly(True)
        font = self.equation_label.font()
        font.setPointSize(14)
        self.equation_label.setFont(font)
        self.result_display = QTextEdit()
        font.setPointSize(12)
        self.result_display.setFont(font)
        self.export_button = QPushButton("Export to Excel",)
        self.export_button.setFixedSize(120,40)
        self.export_button.clicked.connect(self.Export_to_Excel)
        self.result_table = QTableView()
        self.result_table.setModel(self.param_model)
        r_bot_layout.addWidget(self.equation_label,0,0,1,3)
        r_bot_layout.addWidget(self.result_table,1,0,2,2)
        r_bot_layout.addWidget(self.result_display,1,2,Qt.AlignmentFlag.AlignRight)
        r_bot_layout.addWidget(self.export_button,2,2,Qt.AlignmentFlag.AlignRight)
        
        # Message de démarrage        
        self.log.appendLogMessage("Start of session...")
        self.log.appendLogMessage("Session ready. Select a GPC-ONE file to begin.")

    def floryNumberChange(self, value):
        if value < 1:
            self.fit_entry.setValue(1)
            self.log.appendErrorMessage("Number of Flory curve should be at least one.")
        else:
            self.fitFlory()

    def openFileDialog(self) -> None:
        # Open a file dialog
        filename, _ = QFileDialog.getOpenFileName(self, "Select File", "", "All Files (*)")
        # Display the selected file name
        if filename:
            self.log.appendLogMessage(f"Load of file: {filename}")
            self.file_entry.setText(filename)
            # Chargement du fichier
            self.data_GPC.import_file(filename)
            # Affichage des informations de l'échantillons
            self.file_info.clear()
            self.file_info.append(self.data_GPC.__repr__())
            # Fit de la courbe expérimentale
            self.fitFlory()

    def fitFlory(self) -> None:
        logM = self.data_GPC.logM
        w = self.data_GPC.w
        N = self.fit_entry.value()
        # Effectue le fit de la courbe expérimentale
        self.log.appendLogMessage(f"Start of fitting with {N} Flory.")
        try:
            params, error = Flory_fit.fit_N_Flory(logM,w,N)
            self.log.appendLogMessage("End of Fitting with success.")
            self.param_model.setParameters(params, error, N)
            # Affiche la courbe
            self.displayFitFlory()
        except Exception as e:
            self.log.appendErrorMessage(f"Fitting failure. Error : {e}")

    def displayFitFlory(self) -> None:
        # Reset l'affichage des courbes
        self.plot_GPC.clear()
        # Affichage de la courbes de GPC expérimentales
        # pen = pg.mkPen(color=(255, 0, 0))
        self.plot_GPC.scatterPlot(self.data_GPC.logM, self.data_GPC.w,
                                  symbol='o', symbolSize=4, symbolBrush='r', symbolPen=None,
                                  name='GPC Data')
        # Affichage des fits
        w = Flory_fit.get_model_prediction(self.data_GPC.logM,
                                           self.fit_entry.value(),
                                           self.param_model.parameters)
        pen = pg.mkPen(color=(0, 0, 0), width=2)
        self.plot_GPC.plot(self.data_GPC.logM, w[-1,:],pen=pen, name=f"Flory Fit")
        if w.shape[0]>1:
            colors = ('g','m','c','b','y',(200,200,200),(200,200,0),(200,0,200),(200,200,0))
            for i in range(0,w.shape[0]-1):
                pen = pg.mkPen(color=colors[i], style=Qt.PenStyle.DashLine, width=1.5)
                self.plot_GPC.plot(self.data_GPC.logM, w[i,:],pen=pen, name=f"Flory #{i+1}")
        R2 = Flory_fit.r_squared(self.data_GPC.w, w[-1,:])
        self.result_display.clear()
        self.result_display.append("Fit Result :")
        self.result_display.append(f"R² = {R2}")

    def Export_to_Excel(self) -> None:
        file_dialog = QFileDialog(self)
        file_dialog.setFileMode(QFileDialog.FileMode.AnyFile)
        file_dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
        file_dialog.setNameFilter("Excel Files (*.xlsx)")

        if file_dialog.exec() == QFileDialog.DialogCode.Accepted:
            try:
                file_path = file_dialog.selectedFiles()[0]        
                # Create workbook 
                wb = Workbook()
                ws = wb.active
                # Affichage des infos de l'échantillons
                ws.cell(row=1, column=1).value = "Sample informations"
                for i, (key,value) in enumerate(self.data_GPC.dict_info.items(), start=2):
                    ws.cell(row=i, column=1).value = key
                    ws.cell(row=i, column=2).value = value
                # Affichage des résultats
                nb_Flory = self.param_model.nb_Flory
                ws.cell(row=6, column=1).value = "Fitted Flory's Parameters"
                for i, title in enumerate(('m_i','std m_i','Tau_i','std Tau_i','Mn'), start=2):
                    ws.cell(row=7, column=i).value = title
                for i in range(0,nb_Flory):
                    ws.cell(row=8+i, column=1).value = f"Flory #{i+1}"
                    ws.cell(row=8+i, column=2).value = self.param_model.parameters[i]
                    ws.cell(row=8+i, column=3).value = self.param_model.errors[i]
                    ws.cell(row=8+i, column=4).value = self.param_model.parameters[i+nb_Flory]
                    ws.cell(row=8+i, column=5).value = self.param_model.errors[i+nb_Flory]
                    ws.cell(row=8+i, column=6).value = int(np.divide(1, self.param_model.parameters[i+nb_Flory]))
                # Affichage des données de courbes
                w = Flory_fit.get_model_prediction(self.data_GPC.logM,
                                            self.fit_entry.value(),
                                            self.param_model.parameters)
                data = np.concatenate((self.data_GPC.logM.reshape(-1,1),
                                    self.data_GPC.w.reshape(-1,1),
                                    w.T), axis=1)
                titles = ["Log M","W_exp","W_Flory_Fit"]
                if nb_Flory>1:
                    for i, element in enumerate([f"W_Flory_#{k+1}" for k in range(0,nb_Flory)]):
                        titles.insert(2+i,element)
                for i, title in enumerate(titles, start=9):
                    ws.cell(row=1, column=i).value = title
                for r, values in enumerate(data, start=2):
                    for c, value in enumerate(values, start=9):
                        ws.cell(row=r, column=c).value = value
                # Affichage du graph
                chart = ScatterChart()
                chart.x_axis.title = "log(M)"
                chart.y_axis.title = "W"
                x_data = Reference(ws, min_col=9, min_row=2, max_row=1001)
                for i in range(1,np.shape(data)[1]):
                    y_data = Reference(ws, min_col=9+i, min_row=2, max_row=1001)
                    serie = Series(y_data, x_data, title=titles[i])
                    chart.series.append(serie)
                ws.add_chart(chart, "A14")
                # Sauvegarde du fichier
                wb.save(file_path)
                self.log.appendLogMessage("Export to excel finish with success.")
            except Exception as e:
                self.log.appendErrorMessage(f"Error while export to excel : {e}")